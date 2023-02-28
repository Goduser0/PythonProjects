import openpyxl

diquming = '安徽'
# ********************************************


def lookdianchang(diqu):
    excel_diqu = openpyxl.load_workbook('地区分类.xlsx')
    sheet_diqu = excel_diqu['Sheet3']
    data_cols2 = sheet_diqu.iter_cols(2)
    shengfeng = []
    dianchangming = []
    for col in data_cols2:
        for cell in col:
            shengfeng.append(cell.value)
    h = 1
    for sf in shengfeng:
        if sf == str(diqu):
            dcm = str(sheet_diqu.cell(h, 3).value)
            gvm = str(sheet_diqu.cell(h, 4).value)
            gvm = gvm[2:]
            t1 = (dcm, gvm)
            dianchangming.append(t1)
            h += 1
        else:
            h += 1
    print(diqu + '电厂名、风机容量查找 Completed！！！')
    return dianchangming


# print(lookdianchang(diquming))

# ********************************************
# 输入（电厂名，机组容量），找出主轴承、叶轮毂、伺服控制装置、联轴器、叶片的图号


def lookpack(dianchang):
    yilanbiao = openpyxl.load_workbook('设备一览表.xlsx')
    yeji = yilanbiao['国家能源集团业绩']
    name = str(dianchang)
    h = 1
    alltuhao = []

    for cell in yeji["B"]:
        dctuhao = {}
        if cell.value == name:
            dctuhao['电厂名'] = name
            jz = yeji.cell(h, 3).value
            if jz is not None:
                dctuhao['机组容量'] = jz[2:]
                dctuhao['设备名称'] = yeji.cell(h, 4).value
                dctuhao['主轴承装配图号'] = yeji.cell(h, 10).value
                dctuhao['叶轮毂装配图号'] = yeji.cell(h, 11).value
                dctuhao['伺服控制装置图号'] = yeji.cell(h, 12).value
                dctuhao['联轴器图号'] = yeji.cell(h, 13).value
                dctuhao['叶片图号'] = yeji.cell(h, 14).value
                dctuhao['芯轴图号'] = yeji.cell(h, 15).value
            alltuhao.append(dctuhao)
        else:
            pass
        h += 1
    return alltuhao


# lookpack(['国家能源宁夏灵武电厂（原华电）'])
# **************************************
# 依据容量、图号，找出pack的电厂数，风机台数


def look1000pack(tuhao, rl, sb, lj):
    yilanbiao = openpyxl.load_workbook('设备一览表.xlsx')
    ljm = lj
    if rl == '660MW':
        rl1 = '600MW'
    elif rl == '350MW':
        rl1 = '300MW'
    else:
        rl1 = rl
    rl_sheet = yilanbiao[str(rl1)]
    if rl1 == '1000MW':
        h = 0
        for cell in rl_sheet["E"]:
            h += 1
            tuhao_test = []
            if cell.value is not None:
                tuhao_test.append(cell.value)
                for item in tuhao_test:
                    if str(item) == str(tuhao):
                        dcm = str(rl_sheet.cell(h, 6).value)
                        dc_num = 1
                        for dot3 in dcm:
                            if dot3 == '，' or dot3 == ',':
                                dc_num += 1
                        fj_num = rl_sheet.cell(h, 7).value
                        return ljm, tuhao, rl, sb, dc_num, fj_num
            else:
                pass
    else:
        h = 0
        fj_num = 0
        ljm = lj
        dc_num = 0
        for cell in rl_sheet["E"]:
            h += 1
            if cell.value is not None:
                if cell.value == tuhao:
                    dc_num += 1
                    if fj_num == 0:
                        fj_num = rl_sheet.cell(h, 9).value
                        # ljm = rl_sheet.cell(h, 4).value
        return ljm, tuhao, rl, sb, dc_num, fj_num


# print(look1000pack('UZ22000'))
# **************************************


def lookparts(packlx, packbh, rl, fjlx):
    thhz = openpyxl.load_workbook('图号汇总表.xlsx')
    worksheet = thhz[packlx]
    h = 0
    lj_all = []
    if packlx != '叶片':
        for cell in worksheet["A"]:
            h += 1
            if cell.value == str(packbh):
                for h_add in range(18):
                    key = worksheet.cell(h + h_add, 3).value
                    value = worksheet.cell(h + h_add, 4).value
                    lj_all.append((key, value))
            else:
                pass
    else:
        if str(packbh) == worksheet.cell(2, 1).value:
            for h_add in [2, 3]:
                key = worksheet.cell(h_add, 3).value
                value = worksheet.cell(h_add, 4).value
                lj_all.append((key, value))
        if str(packbh) == worksheet.cell(5, 1).value:
            for h_add in [5, 6, 7, 8]:
                key = worksheet.cell(h_add, 3).value
                value = worksheet.cell(h_add, 4).value
                lj_all.append((key, value))
    ylb = openpyxl.load_workbook('设备一览表.xlsx')
    ysj = ylb['易损件统计表 ']
    result = lj_all
    x = -1
    ff = []
    for lj in lj_all:
        x += 1
        h2 = 0
        dc_num = 0
        fj_num = 0
        for tuhao in ysj["D"]:
            h2 += 1
            if lj[1] == tuhao.value:
                end_num = h2
                flag = []
                while not flag:
                    end_num += 1
                    flag = ysj.cell(end_num, 4).value
                    if end_num >= 2767:
                        flag = 1
                start_num = h2
                end_num -= 1
                if rl == '1000MW':
                    for num_h in range(start_num, end_num+1, 3):
                        if isinstance(ysj.cell(num_h, 7).value, int):
                            dc_num += ysj.cell(num_h, 7).value
                        if isinstance(ysj.cell(num_h, 8).value, int):
                            fj_num += ysj.cell(num_h, 8).value
                elif rl == '600MW' or rl == '660MW':
                    for num_h in range(start_num+1, end_num+2, 3):
                        if isinstance(ysj.cell(num_h, 7).value, int):
                            dc_num += ysj.cell(num_h, 7).value
                        if isinstance(ysj.cell(num_h, 8).value, int):
                            fj_num += ysj.cell(num_h, 8).value
                elif rl == '300MW' or rl == '350MW':
                    for num_h in range(start_num+2, end_num+3, 3):
                        if isinstance(ysj.cell(num_h, 7).value, int):
                            dc_num += ysj.cell(num_h, 7).value
                        if isinstance(ysj.cell(num_h, 8).value, int):
                            fj_num += ysj.cell(num_h, 8).value
            else:
                pass
        d = result[x]
        f = (d[0], d[1], rl, fjlx, dc_num, fj_num)
        ff.append(f)
    return ff
