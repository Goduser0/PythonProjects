import openpyxl
from excel_final import lookdianchang
from excel_final import lookpack
from excel_final import look1000pack
from excel_final import lookparts

# diquming = str(input('请输入地区名：'))
for dss in ['宁夏']:
    diquming = dss
    diqu = lookdianchang(diquming)
    dcm = []
    for dcm1 in diqu:
        dcm.append(dcm1[0])
    num_dc = len(dcm)
    print('Total:' + str(num_dc))
    flag_i = 0
    dcm = ['神华宁煤宁东矸石电厂（CFB）']
    print(dcm)
    for dcm1 in dcm:
        flag_i += 1
        excel_diqu = openpyxl.load_workbook(str(diquming) + '.xlsx')
        name = str(dcm1) + '易损件'
        sheet_diqu = excel_diqu.create_sheet(index=0, title=(name + '易损件'))
        pack_tuhao = lookpack(dcm1)

        if pack_tuhao:
            h = 2
            for pack_tuhao_1 in pack_tuhao:
                rl = pack_tuhao_1['机组容量']
                sb = pack_tuhao_1['设备名称']
                zzc_th = pack_tuhao_1['主轴承装配图号']
                yelungu_th = pack_tuhao_1['叶轮毂装配图号']
                sfkz_th = pack_tuhao_1['伺服控制装置图号']
                lzq_th = pack_tuhao_1['联轴器图号']
                yp_th = pack_tuhao_1['叶片图号']
                xz_th = pack_tuhao_1['芯轴图号']
                if zzc_th is not None and rl is not None:
                    sheet_diqu.cell(1, 1).value = '序号'
                    sheet_diqu.cell(1, 2).value = '装配部件'
                    sheet_diqu.cell(1, 3).value = '图号'
                    sheet_diqu.cell(1, 4).value = '机组容量'
                    sheet_diqu.cell(1, 5).value = '风机类型'
                    sheet_diqu.cell(1, 6).value = '适用电厂数'
                    sheet_diqu.cell(1, 7).value = '风机台数'
                    for line in [2, 3, 4, 5, 6, 7]:
                        wait_in = look1000pack(zzc_th, rl, sb, '主轴承装配')
                        if wait_in and wait_in[line - 2]:
                            sheet_diqu.cell(h, line).value = wait_in[line - 2]
                    parts = lookparts('主轴承装配', zzc_th, rl, sb)
                    for part in parts:
                        h += 1
                        if part[0] == '芯轴' or part[0] is None:
                            h -= 1
                            continue
                        else:
                            sheet_diqu.cell(h, 2).value = part[0]
                            sheet_diqu.cell(h, 3).value = part[1]
                            sheet_diqu.cell(h, 4).value = part[2]
                            sheet_diqu.cell(h, 5).value = part[3]
                            sheet_diqu.cell(h, 6).value = part[4]
                            sheet_diqu.cell(h, 7).value = part[5]

                    for line in [2, 3, 4, 5, 6, 7]:
                        wait_in = look1000pack(yelungu_th, rl, sb, '叶轮毂装配')
                        if wait_in and wait_in[line - 2]:
                            sheet_diqu.cell(h, line).value = wait_in[line - 2]
                    parts = lookparts('叶轮毂装配', yelungu_th, rl, sb)
                    for part in parts:
                        h += 1
                        sheet_diqu.cell(h, 2).value = part[0]
                        sheet_diqu.cell(h, 3).value = part[1]
                        sheet_diqu.cell(h, 4).value = part[2]
                        sheet_diqu.cell(h, 5).value = part[3]
                        sheet_diqu.cell(h, 6).value = part[4]
                        sheet_diqu.cell(h, 7).value = part[5]

                    for line in [2, 3, 4, 5, 6, 7]:
                        wait_in = look1000pack(sfkz_th, rl, sb, '伺服控制装置')
                        if wait_in and wait_in[line - 2]:
                            sheet_diqu.cell(h, line).value = wait_in[line - 2]
                    parts = lookparts('叶片', yp_th, rl, sb)
                    for part in parts:
                        h += 1
                        sheet_diqu.cell(h, 2).value = part[0]
                        sheet_diqu.cell(h, 3).value = part[1]
                        sheet_diqu.cell(h, 4).value = part[2]
                        sheet_diqu.cell(h, 5).value = part[3]
                        sheet_diqu.cell(h, 6).value = part[4]
                        sheet_diqu.cell(h, 7).value = part[5]
                    h += 1

                    for line in [2, 3, 4, 5, 6, 7]:
                        wait_in = look1000pack(lzq_th, rl, sb, '联轴器')
                        if wait_in and wait_in[line - 2]:
                            sheet_diqu.cell(h, line).value = wait_in[line - 2]
                    h += 1

                    for line in [2, 3, 4, 5, 6, 7]:
                        wait_in = look1000pack(yp_th, rl, sb, '叶片')
                        if wait_in and wait_in[line - 2]:
                            sheet_diqu.cell(h, line).value = wait_in[line - 2]
                    h += 1
                    for line in [2, 3, 4, 5, 6, 7]:
                        wait_in = look1000pack(xz_th, rl, sb, '芯轴')
                        if wait_in and wait_in[line - 2]:
                            sheet_diqu.cell(h, line).value = wait_in[line - 2]
                    h += 1
        excel_diqu.save(str(diquming) + '.xlsx')
        print(dcm1 + 'Search Completed---' + str(flag_i) + '/' + str(num_dc))
    print('******** ALL DONE ********')
