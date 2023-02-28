import openpyxl
zhuangji = openpyxl.load_workbook('浙能装机（转子内容）.xlsx')
sheet1 = zhuangji['Sheet1']
lbj = openpyxl.load_workbook('零部件.xlsx')

worksheet = lbj['Sheet1']
worksheet.cell(1, 1).value = '台'
act_sym = 'A' + str(1) + ':' + 'F' + str(1)
worksheet.merge_cells(act_sym)

lbj.save('零部件.xlsx')
