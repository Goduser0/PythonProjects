import openpyxl
from my_function import lookparts
from openpyxl.styles import Alignment

zhuangji = openpyxl.load_workbook('浙能装机（转子内容）.xlsx')
sheet1 = zhuangji['Sheet1']
lbj = openpyxl.load_workbook('零部件.xlsx')

detail = []
for hang in range(2, 30):
    user_name = str(sheet1.cell(hang, 2).value)
    size = str(sheet1.cell(hang, 3).value)
    machine_name = str(sheet1.cell(hang, 4).value)
    number = str(sheet1.cell(hang, 5).value)
    NO_zzc = str(sheet1.cell(hang, 9).value)
    NO_1 = str(sheet1.cell(hang, 10).value)
    NO_2 = str(sheet1.cell(hang, 11).value)
    NO_sf = str(sheet1.cell(hang, 12).value)
    NO_lzq = str(sheet1.cell(hang, 13).value)
    NO_yp = str(sheet1.cell(hang, 14).value)
    NO_xz = str(sheet1.cell(hang, 15).value)
    dict_detail = {'用户名称': user_name, '机组容量': size, '设备名称': machine_name, '数量': number, '主轴承装配图号': NO_zzc,
                   'Ⅰ级叶轮毂装配图号': NO_1, 'Ⅱ级叶轮毂装配图号': NO_2, '伺服控制装置': NO_sf, '联轴器': NO_lzq, '叶片': NO_yp,
                   '芯轴': NO_xz}
    detail.append(dict_detail)
# print(detail)

for cell in detail:
    # 工作表初始化
    sheet_name = cell['用户名称']
    if sheet_name in lbj.sheetnames:
        worksheet = lbj[sheet_name]
    else:
        lbj.create_sheet(sheet_name)
        worksheet = lbj[sheet_name]
        A = worksheet.column_dimensions['A']
        A.width = 16
        B = worksheet.column_dimensions['B']
        B.width = 16
        C = worksheet.column_dimensions['C']
        C.width = 5
        D = worksheet.column_dimensions['D']
        D.width = 20
        E = worksheet.column_dimensions['E']
        E.width = 16
        F = worksheet.column_dimensions['F']
        F.width = 14

    # flag_end 为worksheet 的最后一行非零量+1
    flag_begin = 1
    for line_a in worksheet['A']:
        flag_begin += 1
    flag_end = flag_begin

    worksheet.cell(flag_end, 1).value = cell['用户名称'] + ' ' + cell['机组容量'] + ' ' + cell['设备名称'] + ' ' + cell['数量'] + '台'
    act_sym = 'A' + str(flag_end) + ':' + 'F' + str(flag_end)
    worksheet.merge_cells(act_sym)
    flag_end += 1
    worksheet.cell(flag_end, 1).value = '装配图名称'
    worksheet.cell(flag_end, 2).value = '装配图图号'
    worksheet.cell(flag_end, 3).value = '序号'
    worksheet.cell(flag_end, 4).value = '装配部件'
    worksheet.cell(flag_end, 5).value = '装配部件图号'
    worksheet.cell(flag_end, 6).value = '数量（单台）'
    flag = flag_end + 1

    flag_end += 1
    # 主轴承装配
    act_sym = 'A' + str(flag) + ':' + 'A' + str(flag + 17)
    worksheet.merge_cells(act_sym)
    worksheet.cell(flag, 1).value = '主轴承装配'
    act_sym = 'B' + str(flag) + ':' + 'B' + str(flag + 17)
    worksheet.merge_cells(act_sym)
    worksheet.cell(flag, 2).value = cell['主轴承装配图号']

    a = lookparts('主轴承装配', cell['主轴承装配图号'])
    flag_start = flag_end
    for zzc in a:
        worksheet.cell(flag_end, 3).value = flag_end - flag_start + 1
        worksheet.cell(flag_end, 4).value = zzc[0]
        worksheet.cell(flag_end, 5).value = zzc[1]
        worksheet.cell(flag_end, 6).value = zzc[2]
        flag_end += 1

    # Ⅰ级叶轮毂装配图号
    flag += 18
    act_sym = 'A' + str(flag) + ':' + 'A' + str(flag + 17)
    worksheet.merge_cells(act_sym)
    worksheet.cell(flag, 1).value = 'Ⅰ级叶轮毂装配'
    act_sym = 'B' + str(flag) + ':' + 'B' + str(flag + 17)
    worksheet.merge_cells(act_sym)
    worksheet.cell(flag, 2).value = cell['Ⅰ级叶轮毂装配图号']

    b = lookparts('叶轮毂装配', cell['Ⅰ级叶轮毂装配图号'])
    flag_start = flag_end
    for ylg_1 in b:
        worksheet.cell(flag_end, 3).value = flag_end - flag_start + 1
        worksheet.cell(flag_end, 4).value = ylg_1[0]
        worksheet.cell(flag_end, 5).value = ylg_1[1]
        worksheet.cell(flag_end, 6).value = ylg_1[2]
        flag_end += 1

    # Ⅱ级叶轮毂装配图号
    flag += 18
    act_sym = 'A' + str(flag) + ':' + 'A' + str(flag + 17)
    worksheet.merge_cells(act_sym)
    worksheet.cell(flag, 1).value = 'Ⅱ级叶轮毂装配'
    act_sym = 'B' + str(flag) + ':' + 'B' + str(flag + 17)
    worksheet.merge_cells(act_sym)
    worksheet.cell(flag, 2).value = cell['Ⅱ级叶轮毂装配图号']

    c = lookparts('叶轮毂装配', cell['Ⅱ级叶轮毂装配图号'])
    flag_start = flag_end
    for ylg_2 in c:
        worksheet.cell(flag_end, 3).value = flag_end - flag_start + 1
        worksheet.cell(flag_end, 4).value = ylg_2[0]
        worksheet.cell(flag_end, 5).value = ylg_2[1]
        worksheet.cell(flag_end, 6).value = ylg_2[2]
        flag_end += 1
    lbj.save('零部件.xlsx')

    flag_end = flag + 18

    # 伺服控制装置
    worksheet.cell(flag_end, 1).value = '伺服控制装置'
    worksheet.cell(flag_end, 2).value = cell['伺服控制装置']
    flag_end += 1

    # 联轴器
    worksheet.cell(flag_end, 1).value = '联轴器'
    worksheet.cell(flag_end, 2).value = cell['联轴器']
    flag_end += 1
    # 叶片
    worksheet.cell(flag_end, 1).value = '叶片'
    worksheet.cell(flag_end, 2).value = cell['叶片']
    flag_end += 1
    # 芯轴
    worksheet.cell(flag_end, 1).value = '芯轴'
    worksheet.cell(flag_end, 2).value = cell['芯轴']

    # 单元格格式设置
    align = Alignment(horizontal='center', vertical='center')
    for row in range(1, worksheet.max_row+1):
        for column in range(1, worksheet.max_column+1):
            worksheet.cell(row, column).alignment = align
    print(cell['用户名称'] + '-' + cell['设备名称'] + '-DONE!')
    lbj.save('零部件.xlsx')
lbj.remove(lbj['Sheet1'])
lbj.save('零部件.xlsx')
