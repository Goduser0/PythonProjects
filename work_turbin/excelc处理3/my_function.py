import openpyxl

hzb = openpyxl.load_workbook('图号汇总表.xlsx')


def lookparts(packlx, packbh, rl='/', fjlx='/'):
    thhz = openpyxl.load_workbook('图号汇总表.xlsx')
    worksheet = thhz[packlx]
    h = 0
    lj_all = []
    if packlx != '叶片':
        for cell in worksheet["A"]:
            h += 1

            th1 = ''
            th2 = ''
            num_item = 0
            if cell.value:
                for item in cell.value:
                    if item != '/':
                        th1 += item
                    else:
                        th2 = cell.value[num_item + 1:]
                        break
                    num_item += 1

            if th1 == str(packbh) or th2 == str(packbh):
                for h_add in range(18):
                    lj_name = worksheet.cell(h + h_add, 3).value
                    if lj_name is None:
                        lj_name = '/'
                    num_th = worksheet.cell(h + h_add, 4).value
                    if num_th is None:
                        num_th = '/'
                    num_solo = worksheet.cell(h + h_add, 5).value
                    if num_solo is None:
                        num_solo = '/'
                    lj_all.append((lj_name, num_th, num_solo))
            else:
                pass
    else:
        if str(packbh) == worksheet.cell(2, 1).value:
            for h_add in [2, 3]:
                lj_name = worksheet.cell(h + h_add, 3).value
                if lj_name is None:
                    lj_name = '/'
                num_th = worksheet.cell(h + h_add, 4).value
                if num_th is None:
                    num_th = '/'
                num_solo = worksheet.cell(h + h_add, 5).value
                if num_solo is None:
                    num_solo = '/'
                lj_all.append((lj_name, num_th, num_solo))
        if str(packbh) == worksheet.cell(5, 1).value:
            for h_add in [5, 6, 7, 8]:
                lj_name = worksheet.cell(h + h_add, 3).value
                if lj_name is None:
                    lj_name = '/'
                num_th = worksheet.cell(h + h_add, 4).value
                if num_th is None:
                    num_th = '/'
                num_solo = worksheet.cell(h + h_add, 5).value
                if num_solo is None:
                    num_solo = '/'
                lj_all.append((lj_name, num_th, num_solo))

    return lj_all

