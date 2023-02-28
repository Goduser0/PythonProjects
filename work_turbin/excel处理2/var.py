import pickle
import openpyxl


def save_var(v, filename):
    f = open(filename, 'wb')
    pickle.dump(v, f)
    f.close()
    return filename


def load_var(filename):
    f = open(filename, 'wb')
    r = pickle.load(f)
    f.close()
    return r


def del_space(key):
    while key[0] == ' ' or key[-1] == ' ':
        if key[0] == ' ':
            key = key[1:]
        if key[-1] == ' ':
            key = key[:-2]
    return key


yilanbiao = openpyxl.load_workbook('设备一览表-zq.xlsx')
yisunjian = yilanbiao['易损件统计表 ']
yeji = yilanbiao['国家能源集团业绩']


def look(tuhao, diquming):
    diqu_all = {}
    hh = 0

    for yeji_c in yeji["C"]:
        hh += 1
        if yeji_c.value == diquming:

            key_k = yeji.cell(hh, 11).value
            if key_k != '/':
                value_k = [[yeji.cell(hh, 2).value, yeji.cell(hh, 4).value, str(yeji.cell(hh, 6).value),
                            yeji.cell(hh, 8).value, yeji.cell(hh, 5).value]]
                if diqu_all.get(key_k) is not None:
                    old_value = diqu_all[key_k]
                    a = old_value + value_k
                    value_k = a
                    diqu_all[key_k] = value_k
                else:
                    diqu_all[key_k] = value_k

            key_k = yeji.cell(hh, 12).value
            if key_k != '/':
                value_k = [[yeji.cell(hh, 2).value, yeji.cell(hh, 4).value, str(yeji.cell(hh, 6).value),
                            yeji.cell(hh, 8).value, yeji.cell(hh, 5).value]]
                if diqu_all.get(key_k) is not None:
                    old_value = diqu_all[key_k]
                    a = old_value + value_k
                    value_k = a
                    diqu_all[key_k] = value_k
                else:
                    diqu_all[key_k] = value_k

            key_k = yeji.cell(hh, 13).value
            if key_k != '/':
                value_k = [[yeji.cell(hh, 2).value, yeji.cell(hh, 4).value, str(yeji.cell(hh, 6).value),
                            yeji.cell(hh, 8).value, yeji.cell(hh, 5).value]]
                if diqu_all.get(key_k) is not None:
                    old_value = diqu_all[key_k]
                    a = old_value + value_k
                    value_k = a
                    diqu_all[key_k] = value_k
                else:
                    diqu_all[key_k] = value_k

            key_k = yeji.cell(hh, 14).value
            if key_k != '/':
                value_k = [[yeji.cell(hh, 2).value, yeji.cell(hh, 4).value, str(yeji.cell(hh, 6).value),
                            yeji.cell(hh, 8).value, yeji.cell(hh, 5).value]]
                if diqu_all.get(key_k) is not None:
                    old_value = diqu_all[key_k]
                    a = old_value + value_k
                    value_k = a
                    diqu_all[key_k] = value_k
                else:
                    diqu_all[key_k] = value_k

            key_k = yeji.cell(hh, 15).value
            if key_k != '/':
                value_k = [[yeji.cell(hh, 2).value, yeji.cell(hh, 4).value, str(yeji.cell(hh, 6).value),
                            yeji.cell(hh, 8).value, yeji.cell(hh, 5).value]]
                if diqu_all.get(key_k) is not None:
                    old_value = diqu_all[key_k]
                    a = old_value + value_k
                    value_k = a
                    diqu_all[key_k] = value_k
                else:
                    diqu_all[key_k] = value_k

            key_k = yeji.cell(hh, 16).value
            if key_k != '/':
                value_k = [[yeji.cell(hh, 2).value, yeji.cell(hh, 4).value, str(yeji.cell(hh, 6).value),
                            yeji.cell(hh, 8).value, yeji.cell(hh, 5).value]]
                if diqu_all.get(key_k) is not None:
                    old_value = diqu_all[key_k]
                    a = old_value + value_k
                    value_k = a
                    diqu_all[key_k] = value_k
                else:
                    diqu_all[key_k] = value_k

            key_k = yeji.cell(hh, 17).value
            if key_k != '/':
                value_k = [[yeji.cell(hh, 2).value, yeji.cell(hh, 4).value, str(yeji.cell(hh, 6).value),
                            yeji.cell(hh, 8).value, yeji.cell(hh, 5).value]]
                if diqu_all.get(key_k) is not None:
                    old_value = diqu_all[key_k]
                    a = old_value + value_k
                    value_k = a
                    diqu_all[key_k] = value_k
                else:
                    diqu_all[key_k] = value_k
    if diqu_all.get(tuhao) is None:
        return None
    else:
        return diqu_all[tuhao]


def lj_zp(ljth):
    h = 0
    ysjth = {}
    all_end = 2683
    while h <= (all_end - 1):
        zpth = []
        h += 1
        if yisunjian.cell(h, 4).value is not None:
            start = h
            flag1 = True
            while flag1:
                h += 1
                if yisunjian.cell(h, 4).value is not None or h == 2825:
                    flag1 = 0
                    end = h - 1
                    h = end
                    for zp_h in range(start, end, 3):
                        if yisunjian.cell(zp_h, 6).value and yisunjian.cell(zp_h, 6).value != '/':
                            zpth.append(str(yisunjian.cell(zp_h, 5).value))
                            zpth.append(str(yisunjian.cell(zp_h, 6).value))
                        else:
                            zpth.append(str(yisunjian.cell(zp_h, 5).value))
                    key = str(yisunjian.cell(start, 4).value)
                    key = del_space(key)
                    value = zpth
                    ysjth[key] = value
    ljth = str(ljth)
    if ysjth.get(ljth) is None:
        return None
    else:
        return ysjth[ljth]


def mod(oldname):
    if oldname == '国家能源宁夏灵武电厂（原华电）':
        return '灵武'
    if oldname == '宁夏神华国能鸳鸯湖电厂二期':
        return '鸳鸯湖'
    if oldname == '国家能源宁夏大坝电厂（引增合并改造）（原大唐国际）':
        return '大坝'
    if oldname == '神华国华宁东电厂':
        return '国华宁东'
    if oldname == '国家能源大坝电厂四期（原华能）':
        return '大坝'
    if oldname == '国电平罗电厂':
        return '平罗'
    if oldname == '神华国能鸳鸯湖一期（引增合并改造）':
        return '鸳鸯湖'
    if oldname == '国电大武口发电厂':
        return '大武口'
    if oldname == '神华宁煤宁东矸石电厂（CFB）':
        return '宁东矸石'
    if oldname == '国家能源宁夏大坝电厂一、二期改造（原华能）':
        return '大坝'
    if oldname == '国家能源宁夏中卫电厂（原中电投）':
        return '中卫'
    if oldname == '国电大武口热电 (引增合并改造)':
        return '大武口热电'
    if oldname == '河南国电民权电厂':
        return '国电民权'
    if oldname == '国电河南荥阳煤电':
        return '荥阳煤电'
    if oldname == '神华国能焦作电厂':
        return '国能焦作'
    if oldname == '国电河南驻马店热电扩建':
        return '驻马店热电'
    if oldname == '国电河南驻马店#1、2超低排放改造':
        return '驻马店#1、2超低排放改造'
    if oldname == '神皖合肥庐江发电厂':
        return '合肥庐江'
    if oldname == '国电安徽蚌埠电厂（引增合并改造）':
        return '蚌埠'
    if oldname == '国电铜陵#1（引增合并改造）':
        return '铜陵'
    if oldname == '国电宿州发电厂':
        return '宿州'
    if oldname == '神皖安庆皖江#1机组(引增合并改造)':
        return '安庆皖江'
    if oldname == '(神皖集团)安庆电厂二期':
        return '安庆二期'
    else:
        return oldname


def textall(list1, jzrl):
    num = 0
    num_dc = 0
    num_fj = 0
    if list1:
        result = ''
        result1 = ''
        for cell in list1:
            a = mod(cell[0])
            if len(cell) == 5:
                result += (a + cell[3] + cell[4] + '  ')
                num += int(cell[2])
                pl = cell[1]
                pl = pl[-5]
                if str(jzrl) == str(pl):
                    num_dc += 1
                    num_fj += int(cell[2])
                    result1 += (a + cell[3] + cell[4] + '  ')
            else:
                result = cell

        if result1 == '':
            result1 = '/'
        return [result, num, result1, num_dc, num_fj]
    else:
        return ['/', 0, '/']
