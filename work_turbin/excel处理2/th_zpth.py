import openpyxl
from var import look
from var import lj_zp
from var import textall


yilanbiao = openpyxl.load_workbook('设备一览表-zq.xlsx')
num_diquming = input('请输入地区序号：1-宁夏  2-河南  3-安徽: ')
if int(num_diquming) == 1:
    diquming = '宁夏'
elif int(num_diquming) == 2:
    diquming = '河南'
elif int(num_diquming) == 3:
    diquming = '安徽'
else:
    diquming = '宁夏'
    print('地区已默认宁夏')

lxia = openpyxl.load_workbook(diquming + '.xlsx')
yeji = yilanbiao['国家能源集团业绩']

# 此处a为sheet的个数
if diquming == '宁夏':
    a = 8
elif diquming == '河南':
    a = 5
else:
    a = 5

for i in range(a):
    sheet = lxia.worksheets[i]
    h = 0
    for cell_th in sheet["C"]:
        fjs = 0
        tuhao = cell_th.value
        # tuhao = del_space(cell_th.value)
        h += 1
        jzrl = str(sheet.cell(h, 4).value)
        if jzrl[-1] == 'W':
            jzrl = jzrl[-5]

        zpbj = sheet.cell(h, 2).value
        if zpbj == '主轴承装配' or zpbj == 'I级叶轮毂装配' or zpbj == '伺服控制装置' or zpbj == '联轴器' or zpbj == '叶片' or zpbj == '芯轴' or zpbj == 'II级叶轮毂装配' or zpbj == '叶轮毂装配':
            text = look(tuhao, diquming)

        else:
            zpth = lj_zp(tuhao)
            text = []
            if zpth:
                for zpth1 in zpth:
                    if isinstance(zpth1, str):
                        if look(zpth1, diquming):
                            text += look(zpth1, diquming)
                    else:
                        for item in zpth1:
                            text += look(zpth1, diquming)

            else:
                pass

        if text:
            if len(text) != 1:
                while 1:
                    flag = 0
                    xu = None
                    for text1 in text:
                        flag += 1
                        if text1 == '/':
                            xu = flag-1
                            break
                    if xu:
                        text.pop(xu)
                    else:
                        break
        else:
            text = '/'
        text_ready = text

        print(tuhao)
        l1 = len(text_ready)
        if l1 != 1:
            text_nosame = []
            for tr in text_ready:
                if tr not in text_nosame:
                    text_nosame.append(tr)
            text_ready = text_nosame
        print(text_ready)

        if text_ready == '/':
            dcs = 0
        else:
            dcs = len(text_ready)
        text_ok = textall(text_ready, jzrl)[0]
        fjs += int(textall(text_ready, jzrl)[1])
        text_extra = textall(text_ready, jzrl)[2]
        num_dc = textall(text_ready, jzrl)[3]
        num_fj = textall(text_ready, jzrl)[4]
        print(text_ok)
        print(fjs)
        print()
    # dcs fjs text_ok
        sheet.cell(h, 9).value = str(dcs)
        sheet.cell(h, 10).value = str(fjs)
        sheet.cell(h, 11).value = str(text_ok)
        sheet.cell(h, 12).value = str(num_dc)
        sheet.cell(h, 13).value = str(num_fj)
        sheet.cell(h, 14).value = str(text_extra)
        colK = sheet.column_dimensions['K']
        colK.width = 108
        colI = sheet.column_dimensions['I']
        colI.width = 18
        colJ = sheet.column_dimensions['J']
        colJ.width = 18
        colL = sheet.column_dimensions['L']
        colL.width = 30
        colL = sheet.column_dimensions['M']
        colL.width = 30
        colL = sheet.column_dimensions['N']
        colL.width = 60
    sheet.cell(1, 9).value = diquming+'地区电厂数'
    sheet.cell(1, 10).value = diquming+'地区风机数'
    sheet.cell(1, 11).value = diquming+'地区电厂'
    sheet.cell(1, 12).value = diquming+'地区相同机组容量电厂数'
    sheet.cell(1, 13).value = diquming+'地区相同机组容量风机数'
    sheet.cell(1, 14).value = diquming+'地区相同机组容量电厂'
lxia.save(str(diquming) + '.xlsx')
