import os
from pathlib import Path
import openpyxl
from get_FileModifyTime import get_fileModifyTime


def excel_in(fn, mt, pt, sheetname):
    result = openpyxl.load_workbook('result.xlsx')
    sheetname = str(os.path.basename(sheetname))
    result.create_sheet(index=0, title=str(sheetname))
    sheet = result.worksheets[0]
    sheet.cell(1, 1).value = 'file_name'
    sheet.cell(1, 2).value = 'modification_time'
    sheet.cell(1, 3).value = 'file_path'
    n = 2
    for fn1 in fn:
        sheet.cell(n, 1).value = fn1
        n += 1
    n = 2
    for mt1 in mt:
        sheet.cell(n, 2).value = mt1
        n += 1
    n = 2
    for pt1 in pt:
        sheet.cell(n, 3).value = pt1
        n += 1

    result.save('result.xlsx')
    return()


def basiclook(po):
    flag1 = 0
    fn1 = []
    mt1 = []
    pt = []
    for p1 in po.iterdir():
        if not p1.is_file():
            flag1 += 1
        else:
            str_path = str(p1)
            pt.append(str_path)
            fn1.append(os.listdir(po)[flag1])
            str_mtime = str(get_fileModifyTime(p1))
            mt1.append(str_mtime)
            flag1 += 1
    return fn1, mt1, pt


def file_to_excel(file_path):
    fnfinal = []
    mtfinal = []
    pathfinal = []
    for root, dirs, files in os.walk(file_path):
        p = Path(root)
        fnfinal.extend(basiclook(p)[0])
        mtfinal.extend(basiclook(p)[1])
        pathfinal.extend(basiclook(p)[2])
    print('Search Completed...75%')
    excel_in(fnfinal, mtfinal, pathfinal, file_path)
    print('Write into Excel Completed...100%')
    print()
    print('All Done')
file_to_excel('.')